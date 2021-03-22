import openpyxl                                 # import openpyxl module 
from openpyxl.chart import (                    # importing openpyxl for Bar Graph
    Reference,
    Series,
    BarChart3D,
)

# To open the workbook,theFile is created 
theFile = openpyxl.load_workbook('crime.xlsx')
allSheetNames = theFile.sheetnames
# print("All sheet names {} " .format(theFile.sheetnames))
master = theFile['mastersheet']      # to open mastersheet , master is created

print('How Many Input You Want: ')     # Taking how much input user want
b=int(input())
row_data=1

for i in range(0,b):
    cityname=(input('Enter the Area name: '))      # Taking input from user
    for sheet in allSheetNames:                    # itreating through all sheet avilable in my Excel file

        currentSheet = theFile[sheet]              # assiging avilable sheets to variable accoding to the loop

        for row in range(1, currentSheet.max_row + 1):  # itreating through all rows in currentsheet
            #print(row)
            for column in 'ABCDEFGHIJ':             # Here you can add or reduce the columns
                cell_name = "{}{}".format(column, row)
                #print(cell_name)
               
                if currentSheet[cell_name].value == cityname:   # checking if user input is equal to cell value or not
                    r = row                                     # where the input will be equal to cell value it will store that row number.

        for j in range(1,currentSheet.max_column+1):
            master.cell(row=row_data,column=j).value  = currentSheet.cell(row=r,column=j).value   # using this loop to print the data in master sheet 
            # print(row_data)
        row_data=row_data+1

theFile.save('crime.xlsx')


mastersheet = theFile['mastersheet']              # to open mastersheet, mastersheet object is crteated
data = Reference(mastersheet, min_col=2, min_row=1, max_col=3, max_row=4)
titles = Reference(mastersheet, min_col=1, min_row=2, max_row=4)
chart = BarChart3D()
chart.title = "3D Bar Chart"                      # title of the Bar chart
chart.add_data(data=data, titles_from_data=True)
chart.set_categories(titles)
mastersheet.add_chart(chart, "E5")
theFile.save("crime.xlsx")                        # saving the bar graph in mastersheet
